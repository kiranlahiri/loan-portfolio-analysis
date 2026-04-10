"""
interface/components/export.py
--------------------------------
Excel export — multi-sheet workbook for analyst use.

Sheets:
  1. Scenario Comparison  — scenario params, IRR, price targets, CF summary stats
  2. Monthly Cash Flows   — month-by-month projections for all scenarios side by side
  3. Monte Carlo          — IRR distribution summary statistics

Usage:
    from interface.components.export import build_excel
    xlsx_bytes = build_excel(scenario_df, cf_by_scenario, mc, snap, purchase_price)
"""

import io
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Color palette (dark hex fills matching the UI)
# ---------------------------------------------------------------------------
_SCENARIO_FILLS = {
    "Base":   PatternFill("solid", fgColor="1A2A4A"),   # dark blue
    "Stress": PatternFill("solid", fgColor="4A1A1A"),   # dark red
    "Upside": PatternFill("solid", fgColor="1A3A2A"),   # dark green
}
_HEADER_FILL   = PatternFill("solid", fgColor="1C2333")
_SUBHEAD_FILL  = PatternFill("solid", fgColor="252E3F")

_SCENARIO_FONTS = {
    "Base":   Font(bold=True, color="4D94FF"),
    "Stress": Font(bold=True, color="FF4757"),
    "Upside": Font(bold=True, color="2ED573"),
}

_THIN = Side(style="thin", color="334155")
_BORDER = Border(bottom=Side(style="thin", color="334155"))


def _hdr_font(size=10):
    return Font(bold=True, color="B0BEC5", size=size)


def _num_font():
    return Font(name="Courier New", size=9, color="E6EDF3")


def _label_font():
    return Font(size=9, color="90A4AE")


def _apply_header_row(ws, row, values, fill=None):
    """Write a header row with styling."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill
        cell.border = _BORDER


def _fmt_pct(v):
    return round(float(v), 6) if pd.notna(v) else None


def _fmt_dollar(v):
    return round(float(v), 2) if pd.notna(v) else None


# ---------------------------------------------------------------------------
# Sheet 1: Scenario Comparison
# ---------------------------------------------------------------------------

def _write_scenario_sheet(wb: Workbook, scenario_df: pd.DataFrame,
                           cf_by_scenario: dict, purchase_price: float) -> None:
    ws = wb.create_sheet("Scenario Comparison")
    ws.sheet_view.showGridLines = False

    # Title
    ws["A1"] = "Scenario Comparison"
    ws["A1"].font = Font(bold=True, size=13, color="E6EDF3")
    ws["A2"] = f"Purchase price: {purchase_price:.2%} of UPB  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    ws["A2"].font = Font(size=9, color="90A4AE")

    price_cols = [c for c in scenario_df.columns if c.startswith("price_for_")]
    price_labels = [
        "Price @ {}% IRR".format(c.replace("price_for_", "").replace("pct_irr", ""))
        for c in price_cols
    ]

    # Section A: scenario parameters + IRR + prices
    headers_a = ["Scenario", "CDR", "CPR", "Loss Sev.", "IRR"] + price_labels
    _apply_header_row(ws, 4, headers_a, fill=_HEADER_FILL)

    for r_idx, (_, row) in enumerate(scenario_df.iterrows(), 5):
        label = row.get("scenario", "")
        fill  = _SCENARIO_FILLS.get(label, PatternFill("solid", fgColor="1C2333"))
        sfont = _SCENARIO_FONTS.get(label, Font(color="E6EDF3", size=9))

        vals = [
            label,
            _fmt_pct(row["cdr"]),
            _fmt_pct(row["cpr"]),
            _fmt_pct(row["loss_severity"]),
            _fmt_pct(row["irr"]),
        ]
        for col in price_cols:
            vals.append(_fmt_dollar(row.get(col)))

        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.border = _BORDER
            if c_idx == 1:
                cell.font = sfont
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.font = _num_font()
                cell.alignment = Alignment(horizontal="right")
                # Percentage format for CDR/CPR/severity/IRR
                if c_idx <= 5:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "0.0000"

    # Section B: cash flow summary stats
    row_start = 5 + len(scenario_df) + 2
    ws.cell(row=row_start, column=1, value="Cash Flow Summary").font = _hdr_font(11)
    ws.cell(row=row_start, column=1).fill = _SUBHEAD_FILL

    cf_headers = ["Scenario", "Total Interest", "Total Principal", "Total Prepayments",
                  "Total Losses", "Net Cash Flow"]
    _apply_header_row(ws, row_start + 1, cf_headers, fill=_HEADER_FILL)

    for r_idx, (_, row) in enumerate(scenario_df.iterrows(), row_start + 2):
        label = row.get("scenario", "")
        cf = cf_by_scenario.get(label, {})
        fill  = _SCENARIO_FILLS.get(label, PatternFill("solid", fgColor="1C2333"))
        sfont = _SCENARIO_FONTS.get(label, Font(color="E6EDF3", size=9))

        vals = [
            label,
            _fmt_dollar(cf.get("interest",    np.zeros(1)).sum()) if cf else None,
            _fmt_dollar(cf.get("principal",   np.zeros(1)).sum() +
                        cf.get("prepayments", np.zeros(1)).sum()) if cf else None,
            _fmt_dollar(cf.get("prepayments", np.zeros(1)).sum()) if cf else None,
            _fmt_dollar(cf.get("losses",      np.zeros(1)).sum()) if cf else None,
            _fmt_dollar(cf.get("net_cf",      np.zeros(1)).sum()) if cf else None,
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.border = _BORDER
            if c_idx == 1:
                cell.font = sfont
            else:
                cell.font = _num_font()
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'

    # Column widths
    col_widths = [16, 10, 10, 10, 10] + [16] * len(price_cols)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 2: Monthly Cash Flows
# ---------------------------------------------------------------------------

def _write_cashflow_sheet(wb: Workbook, scenario_df: pd.DataFrame,
                          cf_by_scenario: dict) -> None:
    ws = wb.create_sheet("Monthly Cash Flows")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Monthly Cash Flow Projections"
    ws["A1"].font = Font(bold=True, size=13, color="E6EDF3")

    labels = scenario_df["scenario"].tolist()
    cf_fields = ["Interest", "Principal", "Prepayments", "Losses", "Net CF", "Balance SOD"]
    cf_keys   = ["interest", "principal", "prepayments", "losses", "net_cf", "balance_sod"]

    # Header row 1: scenario group labels (merged)
    col = 2
    ws.cell(row=3, column=1, value="Month").font = _hdr_font()
    ws.cell(row=3, column=1).fill = _HEADER_FILL
    for label in labels:
        fill  = _SCENARIO_FILLS.get(label, _HEADER_FILL)
        sfont = _SCENARIO_FONTS.get(label, _hdr_font())
        for i, field in enumerate(cf_fields):
            cell = ws.cell(row=3, column=col + i, value=f"{label} — {field}")
            cell.font = sfont
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right")
            cell.border = _BORDER
        col += len(cf_fields)

    # Determine number of months from first available CF
    n_months = 0
    for label in labels:
        cf = cf_by_scenario.get(label, {})
        if cf:
            n_months = len(cf.get("net_cf", []))
            break

    # Data rows
    for m in range(n_months):
        r = m + 4
        ws.cell(row=r, column=1, value=m + 1).font = _num_font()
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

        col = 2
        for label in labels:
            cf = cf_by_scenario.get(label, {})
            row_fill = _SCENARIO_FILLS.get(label, PatternFill("solid", fgColor="1C2333"))
            for key in cf_keys:
                arr = cf.get(key, np.zeros(n_months))
                val = float(arr[m]) if m < len(arr) else None
                cell = ws.cell(row=r, column=col, value=round(val, 2) if val is not None else None)
                cell.font = _num_font()
                cell.fill = row_fill
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                col += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    col = 2
    for _ in labels:
        for i in range(len(cf_fields)):
            ws.column_dimensions[get_column_letter(col + i)].width = 16
        col += len(cf_fields)


# ---------------------------------------------------------------------------
# Sheet 3: Monte Carlo
# ---------------------------------------------------------------------------

def _write_monte_carlo_sheet(wb: Workbook, mc: dict) -> None:
    ws = wb.create_sheet("Monte Carlo")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Monte Carlo — IRR Distribution Summary"
    ws["A1"].font = Font(bold=True, size=13, color="E6EDF3")
    ws["A2"] = f"Simulations: {len(mc['irrs']):,}  |  CDR and CPR drawn independently from normal distributions"
    ws["A2"].font = Font(size=9, color="90A4AE")

    stats = [
        ("Mean IRR",             mc["mean"]),
        ("Median IRR",           mc["median"]),
        ("Std Dev (IRR)",        mc["std"]),
        ("P5 — 5th Percentile",  mc["p5"]),
        ("P1 — 1st Percentile",  mc["p1"]),
        ("Probability of Loss",  mc["prob_loss"]),
    ]

    _apply_header_row(ws, 4, ["Statistic", "Value"], fill=_HEADER_FILL)

    for r_idx, (label, val) in enumerate(stats, 5):
        fill = PatternFill("solid", fgColor="1C2333")
        lc = ws.cell(row=r_idx, column=1, value=label)
        lc.font = _label_font()
        lc.fill = fill
        lc.border = _BORDER

        vc = ws.cell(row=r_idx, column=2, value=round(float(val), 6))
        vc.font = _num_font()
        vc.fill = fill
        vc.number_format = "0.00%"
        vc.alignment = Alignment(horizontal="right")
        vc.border = _BORDER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    # Limitation note
    ws.cell(row=len(stats) + 7, column=1,
            value="Note: CDR and CPR are drawn independently. "
                  "In practice these are negatively correlated (rising defaults / falling prepayments). "
                  "This model understates correlated tail risk.").font = Font(size=8, color="78909C", italic=True)
    ws.column_dimensions["A"].width = 28


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_excel(
    scenario_df: pd.DataFrame,
    cf_by_scenario: dict,
    mc: dict,
    purchase_price: float,
) -> bytes:
    """
    Build a multi-sheet Excel workbook and return its bytes for download.

    Parameters
    ----------
    scenario_df : pd.DataFrame
        Output of compare_scenarios().
    cf_by_scenario : dict
        {scenario_label: project() output dict} for each scenario.
    mc : dict
        Output of monte_carlo().
    purchase_price : float
        Purchase price as decimal.

    Returns
    -------
    bytes
        Raw .xlsx bytes suitable for st.download_button().
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Dark tab colors
    _write_scenario_sheet(wb, scenario_df, cf_by_scenario, purchase_price)
    wb["Scenario Comparison"].sheet_properties.tabColor = "4D94FF"

    _write_cashflow_sheet(wb, scenario_df, cf_by_scenario)
    wb["Monthly Cash Flows"].sheet_properties.tabColor = "2ED573"

    _write_monte_carlo_sheet(wb, mc)
    wb["Monte Carlo"].sheet_properties.tabColor = "FF4757"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
